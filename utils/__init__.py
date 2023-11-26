import os
from typing import Any, Generator
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd


INPUT_DIR = r"D:\projects\flet_apps\helper-ui\input-files"
DEFAULT_COLUMN_WIDTH = 15


def get_input_files(in_dir: str) -> Generator[str, None, None]:
    return (
        os.path.join(in_dir, filename)
        for filename in os.listdir(in_dir)
        if os.path.isfile(os.path.join(in_dir, filename))
    )


FILES = {os.path.basename(fpath): fpath for fpath in get_input_files(INPUT_DIR)}


def extract_df(excel_file: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    # Load the Excel file into a Pandas DataFrame
    df = pd.read_excel(excel_file)

    # Replace empty strings with None
    df.replace("", None, inplace=True)

    # Identify rows with missing explanations
    missing_explanations = df[df["Explanation"].isnull()]
    return df, missing_explanations


def fetch_incomplete_MCQ(df: pd.DataFrame) -> Generator[dict[str, Any], Any, None]:
    # Iterate through rows with missing explanations and generate/fetch explanations
    for index, row in df.iterrows():
        question = row["Question"]

        # Collect options dynamically by iterating through columns
        options = {
            col_name: value
            for col_name, value in row.items()
            if col_name.startswith("Option") and isinstance(value, str)
        }

        correct_answer = (
            row["Correct Answer"] if isinstance(row["Correct Answer"], str) else ""
        )

        yield {
            "i": index,
            "Question": question,
            "Options": options,
            "Correct Answer": ", ".join(correct_answer.split()).strip(", "),
        }


def set_explanation_detail(
    df: pd.DataFrame,
    index: int,
    explanation: str,
) -> None:
    df["Explanation"] = df["Explanation"].astype(str)
    df.at[index, "Explanation"] = explanation


def setup_workbook(worksheet: Worksheet, df: pd.DataFrame) -> None:
    # Set column widths and text wrapping for each column
    COLUMN_WIDTHS = {
        "Default": DEFAULT_COLUMN_WIDTH,
        "Medium": 25,
        "Large": 40,
    }
    # Set column widths and text wrapping for each column based on column names
    for i, column in enumerate(df.columns, 1):
        if column in ("Question", "Explanation", "Reference"):
            width = COLUMN_WIDTHS["Large"]
        elif "Option" in column:
            width = COLUMN_WIDTHS["Medium"]
        else:
            width = COLUMN_WIDTHS["Default"]

        cell = worksheet.cell(row=1, column=i, value=column)
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        cell.font = Font(bold=True, color="ffffff")
        worksheet.cell(row=1, column=i).fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        worksheet.column_dimensions[chr(64 + i)].width = width

        for j, value in enumerate(df[column], 2):
            cell = worksheet.cell(row=j, column=i, value=value)
            if "Answer" in column:
                worksheet.cell(row=j, column=i).alignment = Alignment(
                    wrapText=True,
                    shrinkToFit=True,
                    horizontal="center",
                    vertical="center",
                )  # Enable text wrapping
            else:
                worksheet.cell(row=j, column=i).alignment = Alignment(
                    wrapText=True,
                    shrinkToFit=True,
                    horizontal="left",
                    vertical="top",
                )  # Enable text wrapping
    # fix top row panes
    worksheet.freeze_panes = "A2"


def save_changes(df: pd.DataFrame, filename: str) -> None:
    try:
        # Export the DataFrame to Excel with text wrapping for 'Column1' using openpyxl engine
        with pd.ExcelWriter(FILES[filename], engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

            # Access the openpyxl workbook and worksheet objects
            workbook = openpyxl.Workbook()
            setup_workbook(workbook.active, df)

        # Save the Excel file
        workbook.save(f"D:\\projects\\flet_apps\helper-ui\\output-files\\{filename}")
    except KeyError as e:
        raise e


def pretty_print_dict(d: dict):
    pretty_dict = ""
    for k, v in d.items():
        pretty_dict += f"\n{k}: "
        if type(v) == dict:
            for value in v:
                pretty_dict += f"\n {value}: {v[value]}"
            else:
                pretty_dict += "\n"
        else:
            pretty_dict += f"{v}\n"

    return pretty_dict.strip()
